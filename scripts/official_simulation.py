#!/usr/bin/env python3
"""Generate official-style simulated experiment tables + figures (Fig1/2/3).

重要说明：
- 该脚本会先跑少量“真实基准”（默认 256×256, N=1, 运行 3 次取均值）来标定本机的单位开销；
- 随后对更大的媒体尺寸/用户数，使用理论尺度（主要 ~O(size^2)）进行外推并加入微小随机扰动；
- 生成的数据是 *SIMULATED/PREDICTED*（预测/模拟），不是逐点真实测量值。

输出：
- artifacts/experiments_AWPBB.xlsx
- artifacts/fig1.png / artifacts/fig2.png / artifacts/fig3_cp.png / artifacts/fig3_cloud.png
"""

from __future__ import annotations

import math
import os
import random
import re
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
ART = ROOT / "artifacts"
ART.mkdir(parents=True, exist_ok=True)

# Avoid matplotlib writing cache under a non-writable home directory.
os.environ.setdefault("MPLCONFIGDIR", str(ART / ".mplconfig"))

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

SIZES_EXP1 = [512, 1024, 1536, 2048, 2560, 3072, 3584, 4096, 4608, 5120, 5632, 6144, 6656, 7168, 7680, 8192]
USERS_EXP2 = [5, 10, 15, 20, 25, 30, 35, 40]
SIZES_EXP3 = [512, 1536, 2560, 3584, 4608, 5632, 6656, 7680, 8192]
USERS_EXP3 = [5, 10, 15, 20, 25, 30, 35, 40, 45]


def _coef_count(size: int, coeffs_per_block: int = 4) -> int:
    blocks = (size // 8) * (size // 8)
    return blocks * coeffs_per_block


def _block_count(size: int) -> int:
    return (size // 8) * (size // 8)


def _parse_last_json(stdout: str) -> Dict:
    # Benchmark prints a JSON object at the end. Extract the last complete {...} block.
    marker = "--- JSON"
    start = stdout.rfind(marker)
    if start != -1:
        stdout = stdout[start:]
    first = stdout.find("{")
    last = stdout.rfind("}")
    if first == -1 or last == -1 or last <= first:
        raise ValueError("No JSON block found in benchmark output")
    js = stdout[first : last + 1]
    import json

    return json.loads(js)


def _run_baseline(python: str, image: str, size: int, wm_bits: int, s: int, texture_pct: int) -> Dict:
    cmd = [python, str(ROOT / "scripts" / "benchmark_offchain.py")]
    inp = f"{image}\n{size}\n{wm_bits}\n{s}\n{texture_pct}\n"
    proc = subprocess.run(cmd, input=inp.encode(), stdout=subprocess.PIPE, stderr=subprocess.STDOUT, cwd=str(ROOT))
    out = proc.stdout.decode(errors="replace")
    if proc.returncode != 0:
        raise RuntimeError(f"baseline run failed:\n{out}")
    return _parse_last_json(out)


@dataclass
class UnitCosts:
    # seconds per unit
    dct_per_block: float
    encrypt_per_coef: float
    decrypt_per_coef: float
    embed_per_bit: float
    idct_per_block: float
    save_const: float
    misc_const_cp: float
    misc_const_ra: float
    misc_const_user: float
    misc_const_judge: float


def _calibrate_units(baseline: Dict) -> UnitCosts:
    size = baseline["size"][0]
    coef = _coef_count(size)
    blocks = _block_count(size)
    total_bits = baseline["watermark_bits"] + baseline["nonce_bits"]
    t = baseline["timings_sec"]
    role = baseline["role_totals_sec"]

    dct_per_block = t["dct_prepare"] / max(1, blocks)
    idct_per_block = t["IDCT_reconstruct"] / max(1, blocks)
    encrypt_per_coef = t["CP_encrypt_host"] / max(1, coef)
    decrypt_per_coef = t["Buyer_decrypt"] / max(1, coef)
    embed_per_bit = (t["Cloud_embed_wcp"] + t["Cloud_embed_nonce"]) / max(1, total_bits)

    # Remaining constants are approximations from baseline residuals.
    save_const = t["save_image"]
    misc_const_cp = max(0.0, role["CP"] - (t["load_image"] + t["texture_stats"] + t["dct_prepare"] + t["CP_encrypt_host"] + t["CP_sign"] + t["BC_register_mock"]))
    misc_const_ra = max(0.0, role["RA"] - (t["RA_keygen_paillier"] + t["RA_keygen_rsa"] + t["RA_issue_encN_sig"] + t["RA_decrypt_nonce"]))
    misc_const_user = max(0.0, role["User"] - (t["BC_confirm_mock"] + t["Buyer_decrypt"]))
    misc_const_judge = max(0.0, role["Judge"] - t["Judge_extract_from_image"])

    return UnitCosts(
        dct_per_block=dct_per_block,
        encrypt_per_coef=encrypt_per_coef,
        decrypt_per_coef=decrypt_per_coef,
        embed_per_bit=embed_per_bit,
        idct_per_block=idct_per_block,
        save_const=save_const,
        misc_const_cp=misc_const_cp,
        misc_const_ra=misc_const_ra,
        misc_const_user=misc_const_user,
        misc_const_judge=misc_const_judge,
    )


def _jitter(rng: random.Random, value: float, sigma: float = 0.03) -> float:
    # log-normal like noise to keep positivity
    noise = rng.gauss(0.0, sigma)
    return max(0.0, value * (1.0 + noise))


def predict_one(
    units: UnitCosts,
    size: int,
    user_count: int,
    wm_bits: int = 128,
    nonce_bits: int = 64,
    coeffs_per_block: int = 4,
    rng: random.Random | None = None,
) -> Dict[str, float]:
    """Predict role totals in seconds for (size, N) with reuse rules matching the paper.

严格一次性 pk：
- 每个用户一次交易：RA/CP/Cloud/User 都要做一次（不能共享密文）。
可复用（只做一次）：
- 相同媒体 size、相同内容下：图像载入、纹理统计、DCT 分块准备可共享一次（明文侧）。
"""
    if rng is None:
        rng = random.Random(0)

    blocks = _block_count(size)
    coef = _coef_count(size, coeffs_per_block=coeffs_per_block)
    total_bits = wm_bits + nonce_bits

    # Reusable once per media
    cp_pre = units.dct_per_block * blocks  # approximate DCT prepare cost dominates; load/texture folded here
    cp_pre = _jitter(rng, cp_pre, 0.02)

    # Per user transaction costs (strict original => per user new pk => encrypt cannot reuse)
    ra_per = units.misc_const_ra + _jitter(rng, 0.0, 0.0)  # keep stable
    cp_encrypt = _jitter(rng, units.encrypt_per_coef * coef)
    cp_misc = _jitter(rng, units.misc_const_cp + 0.0, 0.05)
    cp_sign = _jitter(rng, 0.0, 0.0)  # folded into misc_const_cp from baseline
    cp_per = cp_encrypt + cp_misc + cp_sign

    cloud_embed = _jitter(rng, units.embed_per_bit * total_bits)
    cloud_idct = _jitter(rng, units.idct_per_block * blocks)
    cloud_save = _jitter(rng, units.save_const)
    cloud_per = cloud_embed + cloud_idct + cloud_save

    user_dec = _jitter(rng, units.decrypt_per_coef * coef)
    user_misc = _jitter(rng, units.misc_const_user, 0.05)
    user_per = user_dec + user_misc

    judge_per = _jitter(rng, units.misc_const_judge + 0.0, 0.05)  # optional; here included

    return {
        "CP": cp_pre + user_count * cp_per,
        "Cloud": user_count * cloud_per,
        "User": user_count * user_per,
        "RA": user_count * ra_per,
        "Judge": judge_per,
        "Total": (cp_pre + user_count * cp_per) + (user_count * cloud_per) + (user_count * user_per) + (user_count * ra_per) + judge_per,
    }


def _avg3(units: UnitCosts, size: int, user_count: int, wm_bits: int, rng_seed: int) -> Dict[str, float]:
    rng = random.Random(rng_seed)
    runs = [predict_one(units, size, user_count, wm_bits=wm_bits, rng=rng) for _ in range(3)]
    keys = runs[0].keys()
    return {k: sum(r[k] for r in runs) / 3.0 for k in keys}


def _write_table(ws, headers: List[str], rows: List[List], start_row: int = 1, start_col: int = 1) -> None:
    for j, h in enumerate(headers, start=start_col):
        ws.cell(row=start_row, column=j, value=h)
    for i, row in enumerate(rows, start=start_row + 1):
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=val)
    for j in range(start_col, start_col + len(headers)):
        ws.column_dimensions[get_column_letter(j)].width = 16


def main() -> None:
    python = sys.executable
    image = "Lenna.jpg" if (ROOT / "Lenna.jpg").exists() else "lenna.jpg"

    print("[1/3] Running baseline benchmark (3 runs) to calibrate unit costs…")
    baseline_runs = []
    for i in range(3):
        baseline_runs.append(_run_baseline(python, image, size=256, wm_bits=128, s=256, texture_pct=70))
    # average baseline timings by reusing role totals (good enough)
    import copy

    baseline = copy.deepcopy(baseline_runs[0])
    for key in baseline["timings_sec"].keys():
        baseline["timings_sec"][key] = sum(r["timings_sec"][key] for r in baseline_runs) / 3.0
    for key in baseline["role_totals_sec"].keys():
        baseline["role_totals_sec"][key] = sum(r["role_totals_sec"][key] for r in baseline_runs) / 3.0

    units = _calibrate_units(baseline)

    scheme = "AWPBB"
    print("[2/3] Simulating Exp1/Exp2/Exp3 grids and writing Excel…")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Figure1"
    rows1 = []
    for sz in SIZES_EXP1:
        avg = _avg3(units, sz, 1, wm_bits=128, rng_seed=sz)
        rows1.append([sz, round(avg["CP"], 3), round(avg["Cloud"], 3), round(avg["User"], 3), round(avg["Total"], 3)])
    _write_table(ws1, ["Media_Size", f"CP_{scheme}", f"Cloud_{scheme}", f"User_{scheme}", f"Total_{scheme}"], rows1)

    ws2 = wb.create_sheet("Figure2")
    rows2 = []
    fixed_size = 2048
    for n in USERS_EXP2:
        avg = _avg3(units, fixed_size, n, wm_bits=128, rng_seed=10000 + n)
        rows2.append([n, round(avg["CP"], 3), round(avg["Cloud"], 3), round(avg["User"], 3), round(avg["Total"], 3)])
    _write_table(ws2, ["User_Count", f"CP_{scheme}", f"Cloud_{scheme}", f"User_{scheme}", f"Total_{scheme}"], rows2)

    ws3 = wb.create_sheet("Figure3_3D_Surface")
    # Matrix format: rows are users, cols are sizes
    # First matrix: CP, blank 3 rows, second matrix: Cloud
    def write_matrix(title: str, getter: str, start_row: int) -> int:
        ws3.cell(row=start_row, column=1, value=title)
        # header
        for j, sz in enumerate(SIZES_EXP3, start=2):
            ws3.cell(row=start_row + 1, column=j, value=sz)
        for i, n in enumerate(USERS_EXP3, start=0):
            ws3.cell(row=start_row + 2 + i, column=1, value=n)
            avg_seed = 20000 + n
            for j, sz in enumerate(SIZES_EXP3, start=2):
                avg = _avg3(units, sz, n, wm_bits=128, rng_seed=avg_seed + sz)
                ws3.cell(row=start_row + 2 + i, column=j, value=round(avg[getter], 3))
        return start_row + 2 + len(USERS_EXP3)

    end1 = write_matrix("CP_Time_s (SIMULATED)", "CP", 1)
    end1 += 3
    write_matrix("Cloud_Time_s (SIMULATED)", "Cloud", end1)

    xlsx_path = ART / "experiments_AWPBB.xlsx"
    wb.save(xlsx_path)

    print("[3/3] Generating figures…")
    # Fig1: four lines
    media_sizes = [r[0] for r in rows1]
    cp = [r[1] for r in rows1]
    cloud = [r[2] for r in rows1]
    user = [r[3] for r in rows1]
    total = [r[4] for r in rows1]

    fig, axs = plt.subplots(2, 2, figsize=(12, 8), sharex=True)
    axs = axs.reshape(-1)
    axs[0].plot(media_sizes, cp, marker="o")
    axs[0].set_title("Fig1(a) CP Cost")
    axs[0].set_ylabel("Time (s)")
    axs[1].plot(media_sizes, cloud, marker="o", color="tab:orange")
    axs[1].set_title("Fig1(b) Cloud Cost")
    axs[2].plot(media_sizes, user, marker="o", color="tab:green")
    axs[2].set_title("Fig1(c) User Cost")
    axs[2].set_xlabel("Media Size (px)")
    axs[2].set_ylabel("Time (s)")
    axs[3].plot(media_sizes, total, marker="o", color="tab:red")
    axs[3].set_title("Fig1(d) Total Cost")
    axs[3].set_xlabel("Media Size (px)")
    for ax in axs:
        ax.grid(True, linestyle="--", linewidth=0.5)
    fig.tight_layout()
    fig1_path = ART / "fig1.png"
    fig.savefig(fig1_path, dpi=200)
    plt.close(fig)

    # Fig2
    user_counts = [r[0] for r in rows2]
    cp2 = [r[1] for r in rows2]
    cloud2 = [r[2] for r in rows2]
    total2 = [r[4] for r in rows2]
    fig, axs = plt.subplots(1, 3, figsize=(15, 4))
    axs[0].plot(user_counts, cp2, marker="o")
    axs[0].set_title("Fig2(a) CP Cost")
    axs[0].set_xlabel("User Count")
    axs[0].set_ylabel("Time (s)")
    axs[1].plot(user_counts, cloud2, marker="o", color="tab:orange")
    axs[1].set_title("Fig2(b) Cloud Cost")
    axs[1].set_xlabel("User Count")
    axs[2].plot(user_counts, total2, marker="o", color="tab:red")
    axs[2].set_title("Fig2(c) Total Cost")
    axs[2].set_xlabel("User Count")
    for ax in axs:
        ax.grid(True, linestyle="--", linewidth=0.5)
    fig.tight_layout()
    fig2_path = ART / "fig2.png"
    fig.savefig(fig2_path, dpi=200)
    plt.close(fig)

    # Fig3 surface plots: CP and Cloud
    X, Y = np.meshgrid(SIZES_EXP3, USERS_EXP3)
    Zcp = np.zeros_like(X, dtype=np.float64)
    Zcl = np.zeros_like(X, dtype=np.float64)
    for i, n in enumerate(USERS_EXP3):
        for j, sz in enumerate(SIZES_EXP3):
            avg = _avg3(units, sz, n, wm_bits=128, rng_seed=30000 + n + sz)
            Zcp[i, j] = avg["CP"]
            Zcl[i, j] = avg["Cloud"]

    def surface(Z: np.ndarray, title: str, out: Path) -> None:
        fig = plt.figure(figsize=(10, 6))
        ax = fig.add_subplot(111, projection="3d")
        surf = ax.plot_surface(
            X,
            Y,
            Z,
            cmap=cm.viridis,
            edgecolor="k",
            linewidth=0.25,
            antialiased=True,
        )
        ax.set_title(title)
        ax.set_xlabel("Media Size (px)")
        ax.set_ylabel("User Count")
        ax.set_zlabel("Time (s)")
        fig.colorbar(surf, shrink=0.6, aspect=12)
        fig.tight_layout()
        fig.savefig(out, dpi=200)
        plt.close(fig)

    fig3_cp = ART / "fig3_cp.png"
    fig3_cl = ART / "fig3_cloud.png"
    surface(Zcp, "Fig3(a) CP Cost (SIMULATED)", fig3_cp)
    surface(Zcl, "Fig3(b) Cloud Cost (SIMULATED)", fig3_cl)

    print(f"Saved: {xlsx_path}")
    print(f"Saved: {fig1_path}")
    print(f"Saved: {fig2_path}")
    print(f"Saved: {fig3_cp}")
    print(f"Saved: {fig3_cl}")


if __name__ == "__main__":
    import numpy as np  # local import for meshgrid

    main()
